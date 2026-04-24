#!/usr/bin/env python3
"""Extract application records from an Excel file into individual JSON files.

This script reads a generic Excel file where each row represents one application,
determines column semantics through header analysis, and writes one JSON file
per application to an output directory.

Usage:
    python extract_applications.py <excel_file> [--output-dir <dir>] [--sheet <name>]

Examples:
    python extract_applications.py input_files/30_apps_single_sheet/apps_db_complete.xlsx
    python extract_applications.py apps.xlsx --output-dir output/data/applications
    python extract_applications.py apps.xlsx --sheet "App Details"
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required. Install it with: pip install openpyxl")


# ---------------------------------------------------------------------------
# Column semantics mapping
# ---------------------------------------------------------------------------

# Each entry: canonical_field -> list of patterns (matched case-insensitively)
COLUMN_PATTERNS: dict[str, list[str]] = {
    "app_id": [
        r"^app.?id$",
        r"^id$",
        r"^application.?id$",
        r"^app.?number$",
        r"^application.?number$",
        r"^#$",
        r"^nr$",
        r"^nummer$",
    ],
    "app_name": [
        r"^name$",
        r"^app.?name$",
        r"^application.?name$",
        r"^application$",
        r"^system$",
        r"^anwendung$",
        r"^bezeichnung$",
    ],
    "app_description": [
        r"^description$",
        r"^desc$",
        r"^purpose$",
        r"^function$",
        r"^beschreibung$",
        r"^zweck$",
    ],
    "solution_type": [
        r"solution.?type",
        r"app.?type",
        r"application.?type",
        r"software.?type",
    ],
    "business_criticality": [
        r"criticality",
        r"priority",
        r"business.?impact",
        r"kritikalit",
    ],
    "application_status": [
        r"application.?status",
        r"app.?status",
        r"status",
        r"lifecycle",
    ],
    "decommission_date": [
        r"decommis",
        r"retire",
        r"eol.?date",
        r"end.?of.?life",
        r"sunset",
        r"abschalt",
    ],
    "deployment_type": [
        r"deployment.?type",
        r"hosting",
        r"environment.?type",
        r"deploy",
    ],
    "data_classification": [
        r"data.?classif",
        r"classif",
        r"datenklassif",
    ],
    "business_unit": [
        r"business.?unit",
        r"department",
        r"abteilung",
        r"bereich",
    ],
    "business_capabilities": [
        r"business.?capabilit",
        r"capabilities",
        r"fachliche.?f",
    ],
    "user_count": [
        r"number.?of.?users",
        r"users$",
        r"user.?count",
        r"#.?users",
        r"nutzer",
        r"anzahl.?nutzer",
        r"benutzer",
    ],
    "operating_system": [
        r"operating.?system",
        r"^os$",
        r"platform.?os",
        r"betriebssystem",
    ],
    "programming_language": [
        r"programming.?lang",
        r"language$",
        r"^tech$",
        r"dev.?lang",
        r"sprache",
        r"programmiersprache",
    ],
    "application_server": [
        r"app.?server",
        r"server.?type",
        r"middleware",
        r"application.?server",
    ],
    "application_architecture": [
        r"architecture",
        r"architektur",
        r"app.?arch",
    ],
    "is_containerized": [
        r"container",
        r"docker",
        r"kubernetes",
    ],
    "environment_count": [
        r"number.?of.?env",
        r"env.?count",
        r"#.?env",
        r"umgebungen",
    ],
    "server_instances": [
        r"server.?instance",
        r"physical.?server",
    ],
    "server_count": [
        r"servers$",
        r"#.?servers",
        r"server.?count",
    ],
    "cpu_cores": [
        r"cpu",
        r"cores",
        r"prozessor",
    ],
    "memory_gb": [
        r"memory",
        r"ram",
        r"speicher",
    ],
    "production_environments": [
        r"production.?env",
        r"prod.?env",
    ],
    "ci_cd_present": [
        r"ci.?cd",
        r"pipeline",
        r"continuous",
    ],
    "api_endpoint_count": [
        r"api.?endpoint",
        r"#.?api",
        r"endpoints",
    ],
    "external_interface_count": [
        r"external.?interface",
        r"schnittstellen",
    ],
    "interfaces": [
        r"^interfaces$",
        r"integrations",
        r"connected.?systems",
    ],
    "database_server": [
        r"database.?server",
        r"db.?server",
        r"db.?instance",
    ],
    "database_engine": [
        r"db.?engine",
        r"database$",
        r"^db$",
        r"dbms",
        r"data.?store",
        r"datenbank",
    ],
    "database_storage_gb": [
        r"db.?storage",
        r"storage.?in.?gb",
        r"db.?size",
    ],
    "database_license_required": [
        r"db.?licen",
        r"license.?req",
        r"lizenz",
    ],
    "logging_solution": [
        r"logging",
        r"log.?solution",
        r"protokollierung",
    ],
    "monitoring_tool": [
        r"monitoring",
        r"monitor.?tool",
        r"überwachung",
    ],
    "framework": [
        r"framework",
        r"runtime",
    ],
    "dependencies": [
        r"dependenc",
        r"depends.?on",
        r"upstream",
        r"abhängig",
    ],
    "owner": [
        r"owner",
        r"team",
        r"responsible",
        r"contact",
        r"verantwortlich",
    ],
    "retiring_date": [
        r"retir",
        r"sunset",
    ],
}


def _match_column(header: str) -> str | None:
    """Match a column header to a canonical field name."""
    normalized = header.strip().lower()
    for field, patterns in COLUMN_PATTERNS.items():
        for pattern in patterns:
            if re.search(pattern, normalized):
                return field
    return None


def _sanitize_value(value: Any) -> Any:
    """Convert Excel cell values to JSON-safe types."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value == int(value):
        return int(value)
    return value


def _parse_list_field(value: Any) -> list[str]:
    """Parse a cell value that might contain a comma-separated list."""
    if value is None:
        return []
    text = str(value).strip()
    if not text or text.lower() in ("none", "n/a", "-", ""):
        return []
    return [item.strip() for item in re.split(r"[,;]", text) if item.strip()]


def _auto_parse(value: Any, field_name: str = "") -> Any:
    """Auto-detect lists (comma/semicolon-separated strings) and clean values.

    Fields that contain free text (descriptions, names, suggestions) are never
    split into lists even if they contain commas.
    """
    if value is None:
        return None
    if not isinstance(value, str):
        return value
    text = value.strip()
    if text.lower() in ("none", "n/a", "-", ""):
        return None
    # Never split free-text fields into lists
    never_split = {
        "app_name",
        "app_description",
        "business_unit",
        "data_classification",
        "solution_type",
        "application_status",
        "deployment_type",
        "owner",
        "application_architecture",
        "operating_system",
        "programming_language",
        "application_server",
        "framework",
        "database_engine",
        "logging_solution",
        "monitoring_tool",
    }
    if field_name in never_split:
        return text
    if "," in text or ";" in text:
        items = [item.strip() for item in re.split(r"[,;]", text) if item.strip()]
        if len(items) > 1:
            return items
    return text


def read_excel(
    filepath: Path,
    sheet_name: str | None = None,
) -> tuple[list[str], list[list[Any]]]:
    """Read an Excel file and return headers + data rows."""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            sys.exit(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    print(f"Reading sheet: '{ws.title}' ({ws.max_row - 1} data rows, {ws.max_column} columns)")

    headers = [str(cell.value or "").strip() for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2):
        values = [cell.value for cell in row]
        if all(v is None for v in values):
            continue
        rows.append(values)

    return headers, rows


def map_columns(headers: list[str]) -> dict[int, str]:
    """Map column indices to canonical field names."""
    mapping: dict[int, str] = {}
    used_fields: set[str] = set()

    for idx, header in enumerate(headers):
        if not header:
            continue
        field = _match_column(header)
        if field and field not in used_fields:
            mapping[idx] = field
            used_fields.add(field)
        else:
            safe_key = re.sub(r"[^a-z0-9_]", "_", header.lower().strip())
            safe_key = re.sub(r"_+", "_", safe_key).strip("_")
            mapping[idx] = f"additional:{safe_key}"

    return mapping


def build_application_record(
    row: list[Any],
    column_map: dict[int, str],
    row_index: int,
) -> dict[str, Any]:
    """Build a single application JSON record from a row."""
    record: dict[str, Any] = {}
    additional: dict[str, Any] = {}

    for col_idx, field_name in column_map.items():
        if col_idx >= len(row):
            continue

        raw_value = _sanitize_value(row[col_idx])

        if field_name.startswith("additional:"):
            key = field_name.split(":", 1)[1]
            additional[key] = _auto_parse(raw_value, field_name=key) if isinstance(raw_value, str) else raw_value
        else:
            record[field_name] = _auto_parse(raw_value, field_name=field_name) if isinstance(raw_value, str) else raw_value

    if additional:
        record["additional_attributes"] = additional

    # Ensure app_id exists
    if not record.get("app_id"):
        record["app_id"] = f"APP_{row_index:03d}"

    # Ensure app_name exists
    if not record.get("app_name"):
        record["app_name"] = record["app_id"]

    return record


def write_application_json(
    app: dict[str, Any],
    output_dir: Path,
) -> Path:
    """Write a single application record to a JSON file."""
    app_id = str(app["app_id"]).strip()
    safe_id = re.sub(r"[^a-zA-Z0-9_-]", "_", app_id)
    filepath = output_dir / f"{safe_id}.json"

    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(app, f, indent=2, ensure_ascii=False)

    return filepath


def write_metadata(
    source_file: str,
    column_mapping: dict[str, str],
    app_ids: list[str],
    output_dir: Path,
) -> None:
    """Write extraction metadata to a JSON file."""
    metadata = {
        "source_file": source_file,
        "extraction_date": date.today().isoformat(),
        "total_applications": len(app_ids),
        "column_mapping": column_mapping,
        "application_ids": app_ids,
    }
    filepath = output_dir / "_metadata.json"
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(metadata, f, indent=2, ensure_ascii=False)
    print(f"Metadata written to: {filepath}")


def main() -> None:
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Extract applications from Excel into individual JSON files.",
    )
    parser.add_argument("excel_file", type=Path, help="Path to the Excel file")
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("output/data/applications"),
        help="Directory for output JSON files (default: output/data/applications)",
    )
    parser.add_argument(
        "--sheet",
        type=str,
        default=None,
        help="Sheet name to read (default: active sheet)",
    )
    args = parser.parse_args()

    if not args.excel_file.exists():
        sys.exit(f"File not found: {args.excel_file}")

    # Read Excel
    headers, rows = read_excel(args.excel_file, args.sheet)

    # Map columns
    column_map = map_columns(headers)

    # Print mapping for transparency
    print("\nColumn mapping:")
    readable_mapping: dict[str, str] = {}
    for idx, field in sorted(column_map.items()):
        label = "additional" if field.startswith("additional:") else field
        print(f"  '{headers[idx]}' -> {label}")
        readable_mapping[headers[idx]] = field

    # Create output directory
    args.output_dir.mkdir(parents=True, exist_ok=True)

    # Process rows
    app_ids: list[str] = []
    for row_idx, row in enumerate(rows, start=1):
        app = build_application_record(row, column_map, row_idx)
        filepath = write_application_json(app, args.output_dir)
        app_ids.append(str(app["app_id"]))
        print(f"  [{row_idx}/{len(rows)}] {app['app_id']}: {app.get('app_name', 'N/A')} -> {filepath.name}")

    # Write metadata
    write_metadata(
        source_file=args.excel_file.name,
        column_mapping=readable_mapping,
        app_ids=app_ids,
        output_dir=args.output_dir,
    )

    print(f"\nDone. {len(app_ids)} application(s) extracted to {args.output_dir}/")


if __name__ == "__main__":
    main()
